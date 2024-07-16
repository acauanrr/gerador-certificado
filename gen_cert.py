# import the necessary libraries</pre>

# pip3 install openpyxl
import os
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import img2pdf


def gen_cert_img(details_path,
                 template_path,
                 output_path,
                 font_style,
                 output_path_pdf,
                 pdf_para_rel):

    # loading the details.xlsx workbook
    # and grabbing the active sheet
    obj = openpyxl.load_workbook(details_path)
    sheet = obj.active

    # excel sheet
    for i in range(1, sheet.max_row+1):
        # grab the name of the participant
        name = sheet.cell(row=i, column=1).value
        print("Generating certificate:" + name)

        # grab the id course
        course_id = sheet.cell(row=i, column=2).value

        # grab the course name
        course_name = sheet.cell(row=i, column=3).value

        # grab the number of hours
        number_hours = str(sheet.cell(row=i, column=4).value)

        # grab the date
        date_cert = str(sheet.cell(row=i, column=5).value)

        # style of the font
        font_name = ImageFont.truetype(font_style, 55)
        font_course = ImageFont.truetype(font_style, 40)
        font_numb_hour = ImageFont.truetype(font_style, 35)
        font_date = ImageFont.truetype(font_style, 30)

        # certificate template
        img = Image.open(template_path, mode='r')
        image_width = img.width
        image_height = img.height
        draw = ImageDraw.Draw(img)

        """
        draw the name on the certificate
        """
        text_width = draw.textlength(name, font_name)
        # x,y
        draw.text(
            (
                (image_width - text_width) / 2,
                486
            ),
            name,
            font=font_name,
            fill=(0, 0, 0)
        )

        """
        draw the course name on the certificate
        """
        text_width = draw.textlength(course_name, font_course)
        # x,y
        draw.text(
            (
                (image_width - text_width) / 2,
                720
            ),
            course_name,
            font=font_course,
            fill=(0, 0, 0)
        )

        """
        draw the number of hours on the certificate
        """
        text_width = draw.textlength(number_hours, font_numb_hour)
        # x,y

        draw.text(
            (
                1180,
                858
            ),
            number_hours,
            font=font_numb_hour,
            fill=(0, 0, 0)
        )

        """
        draw the date on the certificate
        """
        text_width = draw.textlength(date_cert, font_date)
        # x,y

        draw.text(
            (
                920,
                1000
            ),
            date_cert,
            font=font_date,
            fill=(0, 0, 0)
        )

        # save the certificate as PNG
        img_pathname = output_path + name + "_" + str(course_id) + ".png"
        img.save(img_pathname)

        pdf_pathname = output_path_pdf + name + "_" + str(course_id) + ".pdf"
        with open(pdf_pathname, "wb") as f:
            f.write(img2pdf.convert(img_pathname))
    
    # Pdf para Relatório (1 file somente)
    imgs = []
    for fname in os.listdir(output_path):
        if not fname.endswith(".png"):
            continue
        path = os.path.join(output_path, fname)
        if os.path.isdir(path):
            continue
        imgs.append(path)
    with open(pdf_para_rel + "pdf-relatorio" + str(course_id) + ".pdf","wb") as f:
        f.write(img2pdf.convert(imgs))


if __name__ == "__main__":

    # style of the font
    font_style = "Playwrite_SK/PlaywriteSK-VariableFont_wght.ttf"

    # template of the certificate
    template_path = 'modelo_certificado_data.png'

    # Excel file containing names of the participants
    details_path = 'list_alunos.xlsx'

    # Output Paths
    output_path = 'certificate_results_png/'
    output_path_pdf = 'certificate_results_pdf/'
    pdf_para_rel = 'pdf_para_rel/'

    gen_cert_img(details_path, template_path,
                 output_path, font_style, output_path_pdf, pdf_para_rel)
